from array import array

from bs4 import BeautifulSoup as bs
from pprint import pprint
import requests, re, os
from urllib.request import  urlretrieve
from openpyxl.drawing.image import Image
from openpyxl import Workbook

wb = Workbook() #엑셀file 껍데기?
sheet1 = wb.active # 시트
sheet1.title = '네이버 웹툰 완결편' #시트명
sheet1.cell(row=1, column=2).value = '제목'
sheet1.cell(row=1, column=3).value = '평점'
sheet1.cell(row=1, column=4).value = '링크'

try:
    if not (os.path.isdir('image')): #image폴더가 없으면
        os.mkdir(os.path.join('image')) #생성
except OSError as e: #에러가 발생한다면
    if e.errno != errno.EEXITS: #존재하는 에러가 아니면
        print("폴더생성실패")
        exit()
url = 'https://comic.naver.com/webtoon/finish.nhn' #네이버 웹툰 완결편 url
html = requests.get(url) #해당하는 url의 html code 모두
soup = bs(html.text, 'html.parser') # BeautifulSoup 함수로 html의text를 분석하기 쉬운 문장으로 만들기
all_data = soup.find('div', {'class':'list_area'}) #  div는 태그, class는 속성, list_area는 속성값
# - div 태그중에 class가 list_area인 놈을 참아서 그 안의 값들을 all_data로 가지고 오라

data_list = all_data.findAll('li') # all_data중에 li로 둘러싼 것들을 list로 가지고 오라 / list 개념을 알아야 하겠지

col = 1
row = 2
''' 
all_data type :  <class 'bs4.element.Tag'>
data_list type :  <class 'bs4.element.ResultSet'>
img_list type :  <class 'bs4.element.Tag'>
'''
for data in data_list: # data_list에서 class?(문장)들을 하나씩 가지고 오기
    img_list = data.find('img') # img태그(속성과 속성값이 없거나 필요없거나 하면 생략, img로도 충분히 가려낼수 있으니)값
    img_src = img_list['src'] # img 태그 안의 src속성에 해당하는 속성값 가져오기 / Beautifulsup으로 분석한 class에서 지원하는 ???
    a_list = data.find('a') # a 태그해당하는
    title = a_list['title'] # a 태그 안의 title속성에 해당하는 속성값 가지고 오기
    title = re.sub('[^0-9a-zA-Zㄱ-힗]', '', title) # 해당하는 문자가 아닌것들(특수문자)은 지우기
    link = "https://comic.naver.com" + a_list['href'] # 기본url를 붙이기
    strong = data.find('strong').text # strong태그의 text값 가져오기 속성값을 제아하지 않는다면 이렇게 간단히 text만 가져오면 됨

   #urlretrieve(img_src, './image/'+title+'.gif') # 해당하는 이미지를 ./image폴더에 제목으로 저장 / 한번만 하면 되므로 주석처리
    img_file = Image('./image/' + title + '.gif') #openpyxl에서 필요한 형식으로
    # pprint(img_file)
    #cell = sheet1.cell(row=row, column=1)
    img_file.anchor = 'A'+str(row) # 해당하는 셀의 좌표
    #pprint('A' + str(col))
    sheet1.add_image(img_file)
    sheet1.cell(row=row, column=2).value = title
    sheet1.cell(row=row, column=3).value = strong
    sheet1.cell(row=row, column=4).value = link
#    col = col + 1
    row = row + 1

wb.save("./webtoon.xlsx")
    #print(title, strong, link)
'''
data2 = data1.findAll('dd')
pprint(data2)
data3 = data2[0].find('span')
print(data3.text)
'''